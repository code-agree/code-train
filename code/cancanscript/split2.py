import openpyxl

def split_text(input_file):
    with open(input_file, 'r', encoding='gb2312') as f:
        lines = f.readlines()
        current_section = []
        workbook = openpyxl.Workbook()

        for line in lines:
            if line.startswith("断面名称"):
                if current_section:
                    section_name = current_section[0].split("：")[1].strip()
                    worksheet = workbook.create_sheet(section_name)
                    worksheet.append([section_name])  # 写入断面名称
                    last_line = False

                    for idx, item in enumerate(current_section[1:]):
                        columns = item.split(',')[:3]  # 获取前三列数据
                        last_line_str = current_section[-1]
                        columns_lst = last_line_str.split(',')[:3]

                        if all(column == '0' for column in columns_lst):
                            last_line = True

                        if all(column == '0' for column in columns[:2]):
                            continue

                        if idx == len(current_section) - 3 and last_line:
                            worksheet.append(columns)
                        else:
                            worksheet.append(columns)

                current_section = [line]

            elif current_section:
                current_section.append(line)

        if current_section:
            section_name = current_section[0].split("：")[1].strip()
            worksheet = workbook.create_sheet(section_name)
            worksheet.append([section_name])  # 写入断面名称

            for item in current_section[1:]:
                columns = item.split(',')[:3]  # 获取前三列数据
                worksheet.append(columns)

        output_file = "output.xlsx"
        workbook.remove(workbook.active)  # 删除默认的Sheet
        workbook.save(output_file)

if __name__ == "__main__":
    input_file = "code-train/code/origin2.txt"
    split_text(input_file)