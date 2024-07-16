import openpyxl
import os

def read_excel_and_write_to_txt(file_path, save_path):
    try:
        # 打开 Excel 文件
        wb = openpyxl.load_workbook(file_path, data_only=True)
        
        # 遍历每个 sheet
        for sheet_name in wb.sheetnames:
            # 获取当前 sheet
            sheet = wb[sheet_name]

            # 获取第一行第一列的值作为文件名
            file_name = str(sheet.cell(row=1, column=1).value)

            # 拼接保存路径和文件名
            txt_file_path = os.path.join(save_path, file_name + '.txt')
            
            # 打开或创建 txt 文件
            with open(txt_file_path, 'w') as txt_file:
                # 写入第一行第一列的值
                txt_file.write(str(sheet.cell(row=1, column=1).value) + '\n')
                
                # 遍历每一行（除了第一行）
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    # 将每一行的前两列写入 txt 文件
                    txt_file.write('\t'.join(str(cell) for cell in row[:2]) + '\n')
            
            print(f"Data from sheet '{sheet_name}' written to {txt_file_path}")
            
    except Exception as e:
        print(f"An error occurred: {e}")

if __name__ == "__main__":
    # 指定 Excel 文件路径
    excel_file_path = 'code-train/code/cancanscript/0606.xlsx'
    
    # 指定保存路径
    save_path = 'code-train/code/cancanscript/result/'
    
    read_excel_and_write_to_txt(excel_file_path, save_path)